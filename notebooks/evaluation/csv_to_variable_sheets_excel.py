import argparse
import math
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TITLE_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F3A66")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SC1_FILL = PatternFill(fill_type="solid", fgColor="DCE6F2")
SC2_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
BEST_BIAS_FILL = PatternFill(fill_type="solid", fgColor="FF0000")
CENTER = Alignment(horizontal="center", vertical="center")


# Columns written to Excel (one "Mean ± Std" string per metric)
DISPLAY_COLUMNS = [
    "Scenario",
    "Dataset",
    "Preupsample",
    "Model",
    "Loss Function",
    "Total Runs",
    "RMSE Mean ± Std",
    "MAE Mean ± Std",
    "Bias Mean ± Std",
    "Pearson Corr Mean ± Std",
]

# Mapping from display column name -> internal numeric mean column (for highlight logic)
METRIC_MEAN_COL = {
    "RMSE Mean ± Std":         "RMSE",
    "MAE Mean ± Std":          "MAE",
    "Bias Mean ± Std":         "Bias",
    "Pearson Corr Mean ± Std": "Pearson Corr",
}


# Highlight criterion per display column
MEAN_METRIC_COLUMNS = {
    "RMSE Mean ± Std":         "min",
    "MAE Mean ± Std":          "min",
    "Bias Mean ± Std":         "abs_min",
    "Pearson Corr Mean ± Std": "max",
}


def _fmt_mean_std(mean: float, std: float) -> str:
    """
    Return 'mean \u00b1 std' string.
    mean : always 4 decimal places  (e.g. '1.4249', '0.0064')
    std  : 4 decimal places when \u2265 5e-5, else scientific 4dp (e.g. '2.8868e-09')
    """
    def _m(v: float) -> str:
        return str(v) if not math.isfinite(v) else f"{v:.4f}"

    def _s(v: float) -> str:
        if not math.isfinite(v):
            return str(v)
        if v == 0.0:
            return "0"
        return f"{v:.4f}" if abs(v) >= 5e-5 else f"{v:.4e}"

    return f"{_m(mean)} \u00b1 {_s(std)}"


def normalize_dataset(value: str) -> str:
    text = str(value).strip().lower()
    if text.startswith("2"):
        return "2 datasets"
    if text.startswith("1"):
        return "1 dataset"
    return str(value).strip()


def infer_scenario_label(raw_scenario: str, dataset: str) -> str:
    scenario = str(raw_scenario).strip().lower()
    if scenario.startswith("scenario1"):
        return "Scenario 1"
    if scenario.startswith("scenario2"):
        return "Scenario 2"

    dataset_text = str(dataset).lower()
    if dataset_text.startswith("2"):
        return "Scenario 1"
    if dataset_text.startswith("1"):
        return "Scenario 2"
    return str(raw_scenario).strip()


def clean_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", name)
    return cleaned[:31] if len(cleaned) > 31 else cleaned


def infer_preupsample_label(value: str) -> str:
    text = str(value).strip().lower()
    if text in {
        "preupsample",
        "yes",
        "true",
        "1",
    }:
        return "Preupsample"

    if text in {
        "no-preupsample",
        "no_preupsample",
        "no preupsample",
        "non-preupsample",
        "no",
        "false",
        "0",
    }:
        return "No Preupsample"

    if "no-preupsample" in text or "no_preupsample" in text or "no preupsample" in text:
        return "No Preupsample"

    return "Preupsample" if "preupsample" in text else "No Preupsample"


def aggregate_table(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()
    work["Dataset"] = work["dataset"].map(normalize_dataset)
    work["Scenario"] = [
        infer_scenario_label(s, d) for s, d in zip(work["scenario"], work["Dataset"])
    ]
    if "preupsample" in work.columns:
        work["Preupsample"] = work["preupsample"].map(infer_preupsample_label)
    elif "filename" in work.columns:
        work["Preupsample"] = work["filename"].map(infer_preupsample_label)
    else:
        work["Preupsample"] = "No Preupsample"

    grouped = (
        work.groupby(
            ["variable", "Scenario", "Dataset", "Preupsample", "model", "loss_function"],
            dropna=False,
            as_index=False,
        )
        .agg(
            total_runs=("run_number", "nunique"),
            rmse_mean=("rmse", "mean"),
            rmse_min=("rmse", "min"),
            rmse_max=("rmse", "max"),
            rmse_std=("rmse", "std"),
            mae_mean=("mae", "mean"),
            mae_min=("mae", "min"),
            mae_max=("mae", "max"),
            mae_std=("mae", "std"),
            bias_mean=("bias", "mean"),
            bias_min=("bias", "min"),
            bias_max=("bias", "max"),
            bias_std=("bias", "std"),
            corr_mean=("corr", "mean"),
            corr_min=("corr", "min"),
            corr_max=("corr", "max"),
            corr_std=("corr", "std"),
        )
    )

    grouped = grouped.rename(
        columns={
            "model": "Model",
            "loss_function": "Loss Function",
            "total_runs": "Total Runs",
            "rmse_mean": "RMSE",
            "rmse_std":  "RMSE Std",
            "mae_mean":  "MAE",
            "mae_std":   "MAE Std",
            "bias_mean": "Bias",
            "bias_std":  "Bias Std",
            "corr_mean": "Pearson Corr",
            "corr_std":  "Pearson Corr Std",
        }
    )

    # fill NaN std (single-run groups) with 0
    for std_col in ["RMSE Std", "MAE Std", "Bias Std", "Pearson Corr Std"]:
        grouped[std_col] = grouped[std_col].fillna(0.0)

    # Build the combined "Mean ± Std" string columns
    grouped["RMSE Mean ± Std"] = [
        _fmt_mean_std(m, s) for m, s in zip(grouped["RMSE"], grouped["RMSE Std"])
    ]
    grouped["MAE Mean ± Std"] = [
        _fmt_mean_std(m, s) for m, s in zip(grouped["MAE"], grouped["MAE Std"])
    ]
    grouped["Bias Mean ± Std"] = [
        _fmt_mean_std(m, s) for m, s in zip(grouped["Bias"], grouped["Bias Std"])
    ]
    grouped["Pearson Corr Mean ± Std"] = [
        _fmt_mean_std(m, s) for m, s in zip(grouped["Pearson Corr"], grouped["Pearson Corr Std"])
    ]

    grouped["scenario_sort"] = grouped["Scenario"].map(
        {"Scenario 1": 1, "Scenario 2": 2}
    ).fillna(99)

    grouped = grouped.sort_values(
        ["variable", "scenario_sort", "Preupsample", "Model", "Loss Function"],
        ascending=[True, True, True, True, True],
    ).drop(columns=["scenario_sort"])

    return grouped


def apply_block_styles(ws, start_row: int, end_row: int, fill: PatternFill) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, len(DISPLAY_COLUMNS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.alignment = CENTER


def highlight_best(ws, start_row: int, end_row: int, numeric: dict) -> None:
    """Bold the best row per metric using pre-extracted numeric means."""
    if start_row > end_row:
        return

    for disp_col, criterion in MEAN_METRIC_COLUMNS.items():
        col_idx = DISPLAY_COLUMNS.index(disp_col) + 1
        mean_col = METRIC_MEAN_COL[disp_col]   # e.g. "RMSE"
        values = numeric[mean_col][start_row - 3 : end_row - 2]  # 0-indexed slice

        if not values:
            continue

        if criterion == "min":
            target = min(values)
            target_rows = [
                start_row + i
                for i, v in enumerate(values)
                if v == target
            ]
        elif criterion == "max":
            target = max(values)
            target_rows = [
                start_row + i
                for i, v in enumerate(values)
                if v == target
            ]
        else:  # abs_min
            target = min(abs(v) for v in values)
            target_rows = [
                start_row + i
                for i, v in enumerate(values)
                if abs(v) == target
            ]

        for row_idx in target_rows:
            ws.cell(row_idx, col_idx).font = Font(bold=True)
            if disp_col == "Bias Mean ± Std":
                ws.cell(row_idx, col_idx).fill = BEST_BIAS_FILL


def autosize_columns(ws) -> None:
    for col_idx in range(1, len(DISPLAY_COLUMNS) + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 52)


def build_workbook(summary: pd.DataFrame, output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    variables = list(summary["variable"].dropna().unique())
    for var in variables:
        var_df = summary[summary["variable"] == var].copy()
        var_table = var_df[DISPLAY_COLUMNS].copy()

        # Keep numeric means for highlight logic (indexed from 0)
        numeric = {
            "RMSE":        var_df["RMSE"].tolist(),
            "MAE":         var_df["MAE"].tolist(),
            "Bias":        var_df["Bias"].tolist(),
            "Pearson Corr": var_df["Pearson Corr"].tolist(),
        }

        ws = wb.create_sheet(title=clean_sheet_name(var))

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DISPLAY_COLUMNS))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"Metrics Comparison - {var}"
        title_cell.fill = TITLE_FILL
        title_cell.font = Font(bold=True, size=16, color="1F3A66")
        title_cell.alignment = CENTER

        # Header row
        for c, col_name in enumerate(DISPLAY_COLUMNS, start=1):
            cell = ws.cell(row=2, column=c, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER

        # Data rows
        current_row = 3
        scenario_ranges = []
        for scenario_name, block in var_table.groupby("Scenario", sort=False):
            block_start = current_row
            for _, row in block.iterrows():
                row_values = {
                    "Scenario":               row["Scenario"],
                    "Dataset":                row["Dataset"],
                    "Preupsample":            row["Preupsample"],
                    "Model":                  row["Model"],
                    "Loss Function":          row["Loss Function"],
                    "Total Runs":             int(row["Total Runs"]),
                    "RMSE Mean ± Std":        str(row["RMSE Mean ± Std"]),
                    "MAE Mean ± Std":         str(row["MAE Mean ± Std"]),
                    "Bias Mean ± Std":        str(row["Bias Mean ± Std"]),
                    "Pearson Corr Mean ± Std": str(row["Pearson Corr Mean ± Std"]),
                }
                for col_idx, col_name in enumerate(DISPLAY_COLUMNS, start=1):
                    cell = ws.cell(current_row, col_idx, row_values[col_name])
                    cell.alignment = CENTER
                current_row += 1
            block_end = current_row - 1
            scenario_ranges.append((scenario_name, block_start, block_end))

        # Merge scenario labels and apply fills/highlights per scenario block
        for scenario_name, start_row, end_row in scenario_ranges:
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                ws.cell(start_row, 1).alignment = CENTER

            block_fill = SC1_FILL if scenario_name == "Scenario 1" else SC2_FILL
            apply_block_styles(ws, start_row, end_row, block_fill)
            highlight_best(ws, start_row, end_row, numeric)

        # Total Runs column: integer format
        total_runs_col = DISPLAY_COLUMNS.index("Total Runs") + 1
        for r in range(3, ws.max_row + 1):
            ws.cell(r, total_runs_col).number_format = "0"

        ws.freeze_panes = "A3"
        autosize_columns(ws)

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert parsed notebook metrics CSV into an Excel workbook with one sheet per variable."
    )
    parser.add_argument(
        "--input",
        default="parsed_notebook_metrics.csv",
        help="Path to input CSV file",
    )
    parser.add_argument(
        "--output",
        default="metrics_comparison_by_variable.xlsx",
        help="Path to output Excel (.xlsx) file",
    )
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    output_path = Path(args.output).resolve()

    if not input_path.exists():
        raise FileNotFoundError(f"Input CSV not found: {input_path}")

    df = pd.read_csv(input_path)
    required = {
        "scenario",
        "dataset",
        "model",
        "loss_function",
        "variable",
        "rmse",
        "mae",
        "bias",
        "corr",
    }
    missing = required.difference(df.columns)
    if missing:
        raise ValueError(f"Missing required column(s): {', '.join(sorted(missing))}")

    clean = df.dropna(subset=["variable", "rmse", "mae", "bias", "corr"]).copy()
    summary = aggregate_table(clean)

    if summary.empty:
        raise ValueError("No valid metric rows were found after filtering input data.")

    build_workbook(summary, output_path)
    print(f"Excel file created: {output_path}")


if __name__ == "__main__":
    main()
