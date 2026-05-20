# -*- coding: utf-8 -*-
"""
generate_metrics_excel.py
=========================
Reads  parsed_notebook_metrics.csv  and writes
metrics_comparison_by_variable.xlsx.

Output layout (mirrors csv_to_variable_sheets_excel.py):
  - One sheet per climate variable
  - Styled title row + dark-blue header row
  - Scenario colour blocks (SC1 = blue tint, SC2 = green tint)
  - Scenario column merged for each block
  - Best-value cells bolded; best Bias cell highlighted in red

Precision rules:
  mean  -> always 4 decimal places   e.g. 1.4249 / 0.0064
  std   -> 4 decimal places when >= 5e-5,
           else scientific 4dp       e.g. 2.8868e-09

Run:
    python generate_metrics_excel.py
"""

import math
import os
import re
import sys

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── paths ─────────────────────────────────────────────────────────────────────
BASE      = os.path.dirname(os.path.abspath(__file__))
CSV_PATH  = os.path.join(BASE, "parsed_notebook_metrics.csv")
XLSX_PATH = os.path.join(BASE, "metrics_comparison_by_variable.xlsx")


# ── styles ────────────────────────────────────────────────────────────────────
TITLE_FILL     = PatternFill(fill_type="solid", fgColor="D9E1F2")
HEADER_FILL    = PatternFill(fill_type="solid", fgColor="1F3A66")
HEADER_FONT    = Font(color="FFFFFF", bold=True)
SC1_FILL       = PatternFill(fill_type="solid", fgColor="DCE6F2")
SC2_FILL       = PatternFill(fill_type="solid", fgColor="E2F0D9")
BEST_BIAS_FILL = PatternFill(fill_type="solid", fgColor="FF0000")
CENTER         = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ── display columns ───────────────────────────────────────────────────────────
DISPLAY_COLUMNS = [
    "Scenario",
    "Dataset",
    "Preupsample",
    "Model",
    "Loss Function",
    "Total Runs",
    "MAE Mean \u00b1 Std",
    "RMSE Mean \u00b1 Std",
    "Bias Mean \u00b1 Std",
    "Corr Mean \u00b1 Std",
    "Skill Mean \u00b1 Std",
    "Baseline RMSE Mean \u00b1 Std",
]

# Highlight criterion: display-col -> (numeric_col, criterion)
HIGHLIGHT_CRITERIA = {
    "MAE Mean \u00b1 Std":          ("mae_mean",   "min"),
    "RMSE Mean \u00b1 Std":         ("rmse_mean",  "min"),
    "Bias Mean \u00b1 Std":         ("bias_mean",  "abs_min"),
    "Corr Mean \u00b1 Std":         ("corr_mean",  "max"),
    "Skill Mean \u00b1 Std":        ("skill_mean", "max"),
}


# ── precision formatters ─────────────────────────────────────────────────────
# mean -> always 4 decimal places
# std  -> 4 decimal places when >= 5e-5, else scientific 4dp (e.g. 2.8868e-09)

def _fmt_mean(v: float) -> str:
    """Format the mean: always 4 decimal places."""
    if not math.isfinite(v):
        return str(v)
    return f"{v:.4f}"


def _fmt_std(std: float) -> str:
    """
    Format the std:
      - 0          -> '0'
      - abs >= 5e-5 -> 4 decimal places  (e.g. '0.0001')
      - abs <  5e-5 -> scientific 4dp    (e.g. '2.8868e-09')
    """
    if not math.isfinite(std):
        return str(std)
    if std == 0.0:
        return "0"
    if abs(std) >= 5e-5:
        return f"{std:.4f}"
    return f"{std:.4e}"


def _fmt_pair(mean: float, std: float) -> str:
    """Return 'mean +/- std' string."""
    return f"{_fmt_mean(mean)} \u00b1 {_fmt_std(std)}"


def mean_pm_std(values) -> str:
    """Return 'mean +/- std' string from an array of values."""
    arr = np.asarray(values, dtype=float)
    arr = arr[np.isfinite(arr)]
    if len(arr) == 0:
        return "N/A"
    m = float(np.mean(arr))
    s = float(np.std(arr, ddof=1) if len(arr) > 1 else 0.0)
    return _fmt_pair(m, s)


# ── label helpers ─────────────────────────────────────────────────────────────

def _scenario_label(raw: str, dataset: str) -> str:
    text = str(raw).strip().lower()
    if text.startswith("scenario1"):
        return "Scenario 1"
    if text.startswith("scenario2"):
        return "Scenario 2"
    ds = str(dataset).lower()
    if ds.startswith("2"):
        return "Scenario 1"
    if ds.startswith("1"):
        return "Scenario 2"
    return str(raw).strip()


def _dataset_label(val: str) -> str:
    text = str(val).strip().lower()
    if text.startswith("2"):
        return "2 datasets"
    if text.startswith("1"):
        return "1 dataset"
    return str(val).strip()


def _preupsample_label(val: str) -> str:
    text = str(val).strip().lower()
    if text in {"preupsample", "yes", "true", "1"}:
        return "Preupsample"
    if "no" in text or text in {"false", "0"}:
        return "No Preupsample"
    return "Preupsample" if "preupsample" in text else "No Preupsample"


def _clean_sheet(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", name)
    return cleaned[:31]


# ── aggregation ───────────────────────────────────────────────────────────────

def aggregate(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()
    work.columns = [c.strip().lower() for c in work.columns]

    for col in ["rmse", "mae", "bias", "corr", "skill", "baseline_rmse"]:
        if col in work.columns:
            work[col] = pd.to_numeric(work[col], errors="coerce")

    work["Dataset"]     = work["dataset"].map(_dataset_label)
    work["Scenario"]    = [_scenario_label(s, d) for s, d in
                            zip(work["scenario"], work["Dataset"])]
    work["Preupsample"] = work["preupsample"].map(_preupsample_label) \
        if "preupsample" in work.columns else "No Preupsample"

    model_col = "model"
    loss_col  = "loss_function" if "loss_function" in work.columns else None
    run_col   = "run_number"    if "run_number"    in work.columns else None

    group_by = ["variable", "Scenario", "Dataset", "Preupsample", model_col]
    if loss_col:
        group_by.append(loss_col)

    agg_dict = {
        "rmse_mean":  ("rmse",  "mean"), "rmse_std":  ("rmse",  "std"),
        "mae_mean":   ("mae",   "mean"), "mae_std":   ("mae",   "std"),
        "bias_mean":  ("bias",  "mean"), "bias_std":  ("bias",  "std"),
        "corr_mean":  ("corr",  "mean"), "corr_std":  ("corr",  "std"),
        "skill_mean": ("skill", "mean"), "skill_std": ("skill", "std"),
    }
    if "baseline_rmse" in work.columns:
        agg_dict["baseline_mean"] = ("baseline_rmse", "mean")
        agg_dict["baseline_std"]  = ("baseline_rmse", "std")
    if run_col:
        agg_dict["total_runs"] = (run_col, "nunique")

    grouped = (
        work.groupby(group_by, dropna=False, as_index=False)
        .agg(**agg_dict)
    )

    for c in ["rmse_std", "mae_std", "bias_std", "corr_std", "skill_std", "baseline_std"]:
        if c in grouped.columns:
            grouped[c] = grouped[c].fillna(0.0)

    # Build formatted string columns
    grouped["MAE Mean \u00b1 Std"] = [
        _fmt_pair(m, s) for m, s in zip(grouped["mae_mean"], grouped["mae_std"])
    ]
    grouped["RMSE Mean \u00b1 Std"] = [
        _fmt_pair(m, s) for m, s in zip(grouped["rmse_mean"], grouped["rmse_std"])
    ]
    grouped["Bias Mean \u00b1 Std"] = [
        _fmt_pair(m, s) for m, s in zip(grouped["bias_mean"], grouped["bias_std"])
    ]
    grouped["Corr Mean \u00b1 Std"] = [
        _fmt_pair(m, s) for m, s in zip(grouped["corr_mean"], grouped["corr_std"])
    ]
    grouped["Skill Mean \u00b1 Std"] = [
        _fmt_pair(m, s) for m, s in zip(grouped["skill_mean"], grouped["skill_std"])
    ]
    if "baseline_mean" in grouped.columns:
        grouped["Baseline RMSE Mean \u00b1 Std"] = [
            _fmt_pair(m, s)
            for m, s in zip(grouped["baseline_mean"], grouped["baseline_std"])
        ]

    grouped = grouped.rename(columns={
        model_col: "Model",
        **({"loss_function": "Loss Function"} if loss_col else {}),
        **({"total_runs": "Total Runs"}       if run_col  else {}),
    })
    if "Total Runs" not in grouped.columns:
        grouped["Total Runs"] = 1
    if "Loss Function" not in grouped.columns:
        grouped["Loss Function"] = "n/a"

    grouped["_sc_sort"] = grouped["Scenario"].map(
        {"Scenario 1": 1, "Scenario 2": 2}
    ).fillna(99)
    grouped = grouped.sort_values(
        ["variable", "_sc_sort", "Preupsample", "Model", "Loss Function"],
        ascending=True,
    ).drop(columns=["_sc_sort"])

    return grouped


# ── workbook builder ──────────────────────────────────────────────────────────

def _autosize(ws, n_cols: int) -> None:
    for col_idx in range(1, n_cols + 1):
        letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0)
            for c in ws[letter]
        )
        ws.column_dimensions[letter].width = min(max(max_len + 3, 12), 52)


def _apply_block_fill(ws, start_row: int, end_row: int,
                      fill: PatternFill, n_cols: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill      = fill
            cell.alignment = CENTER


def _highlight_best(ws, start_row: int, end_row: int,
                    numeric: dict, disp_cols: list) -> None:
    """Bold best-value cells; red-fill best Bias cell."""
    if start_row > end_row:
        return
    for disp_col, (num_col, criterion) in HIGHLIGHT_CRITERIA.items():
        if disp_col not in disp_cols or num_col not in numeric:
            continue
        col_idx = disp_cols.index(disp_col) + 1
        # slice the numeric values that belong to this scenario block
        values = numeric[num_col][start_row - 3 : end_row - 2]
        if not values:
            continue
        finite = [v for v in values if math.isfinite(v)]
        if not finite:
            continue

        if criterion == "min":
            target = min(finite)
            target_rows = [start_row + i for i, v in enumerate(values) if v == target]
        elif criterion == "max":
            target = max(finite)
            target_rows = [start_row + i for i, v in enumerate(values) if v == target]
        else:  # abs_min
            target = min(abs(v) for v in finite)
            target_rows = [start_row + i for i, v in enumerate(values) if abs(v) == target]

        for row_idx in target_rows:
            ws.cell(row_idx, col_idx).font = Font(bold=True)
            if disp_col == "Bias Mean \u00b1 Std":
                ws.cell(row_idx, col_idx).fill = BEST_BIAS_FILL


def build_workbook(summary: pd.DataFrame) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # only keep display columns that actually exist in the dataframe
    disp_cols = [c for c in DISPLAY_COLUMNS if c in summary.columns]
    n_cols = len(disp_cols)

    variables = list(summary["variable"].dropna().unique())

    for var in variables:
        var_df = summary[summary["variable"] == var].copy().reset_index(drop=True)

        # numeric means indexed 0..N-1 (same row order as var_df)
        numeric = {
            num_col: var_df[num_col].tolist() if num_col in var_df.columns else []
            for _, (num_col, _) in HIGHLIGHT_CRITERIA.items()
        }

        ws = wb.create_sheet(title=_clean_sheet(var))

        # ── row 1: title ──────────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        tc = ws.cell(row=1, column=1)
        tc.value     = f"Metrics Comparison - {var}"
        tc.fill      = TITLE_FILL
        tc.font      = Font(bold=True, size=14, color="1F3A66")
        tc.alignment = CENTER

        # ── row 2: headers ────────────────────────────────────────────────
        for c, col_name in enumerate(disp_cols, start=1):
            cell = ws.cell(row=2, column=c, value=col_name)
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = CENTER

        # ── data rows ─────────────────────────────────────────────────────
        current_row     = 3
        scenario_ranges = []

        for scenario_name, block in var_df.groupby("Scenario", sort=False):
            block_start = current_row
            for _, row in block.iterrows():
                for col_idx, col_name in enumerate(disp_cols, start=1):
                    val = row.get(col_name, "")
                    if col_name == "Total Runs" and val != "":
                        try:
                            val = int(val)
                        except (ValueError, TypeError):
                            pass
                    cell = ws.cell(current_row, col_idx, val)
                    cell.alignment = CENTER
                current_row += 1
            block_end = current_row - 1
            scenario_ranges.append((scenario_name, block_start, block_end))

        # ── style each scenario block ─────────────────────────────────────
        for scenario_name, start_row, end_row in scenario_ranges:
            block_fill = SC1_FILL if scenario_name == "Scenario 1" else SC2_FILL
            _apply_block_fill(ws, start_row, end_row, block_fill, n_cols)

            # merge scenario label column
            if "Scenario" in disp_cols:
                sc_col = disp_cols.index("Scenario") + 1
                if end_row > start_row:
                    ws.merge_cells(
                        start_row=start_row, start_column=sc_col,
                        end_row=end_row,     end_column=sc_col,
                    )
                ws.cell(start_row, sc_col).alignment = CENTER

            _highlight_best(ws, start_row, end_row, numeric, disp_cols)

        # ── Total Runs: integer format ────────────────────────────────────
        if "Total Runs" in disp_cols:
            tr_col = disp_cols.index("Total Runs") + 1
            for r in range(3, ws.max_row + 1):
                ws.cell(r, tr_col).number_format = "0"

        ws.freeze_panes = "A3"
        _autosize(ws, n_cols)

    wb.save(XLSX_PATH)
    print(f"[OK] Excel written -> {XLSX_PATH}")
    print(f"  Variables : {len(variables)}")
    print(f"  Rows total: {len(summary)}")


# ── entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    if not os.path.isfile(CSV_PATH):
        print(f"[ERROR] CSV not found: {CSV_PATH}", file=sys.stderr)
        sys.exit(1)

    df = pd.read_csv(CSV_PATH)
    required = {"scenario", "dataset", "model", "variable", "rmse", "mae", "bias", "corr"}
    missing  = required.difference({c.lower() for c in df.columns})
    if missing:
        print(f"[ERROR] Missing columns: {', '.join(sorted(missing))}", file=sys.stderr)
        sys.exit(1)

    summary = aggregate(df)
    if summary.empty:
        print("[ERROR] No valid rows after aggregation.", file=sys.stderr)
        sys.exit(1)

    build_workbook(summary)


if __name__ == "__main__":
    main()
